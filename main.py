# main.py final con comando /subirbackup y subida secuencial a las 18:30 hora Perú
import os
import re
import json
import logging
import asyncio
import os.path
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from datetime import datetime
import nest_asyncio
from dotenv import load_dotenv
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler, CallbackQueryHandler,
    ContextTypes, filters
)
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google.oauth2 import service_account
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from pytz import timezone
from PIL import Image as PILImage

# Zona horaria de Lima (UTC-5)
LIMA_TZ = timezone("America/Lima")

# Cargar variables de entorno
load_dotenv()  # <--- Coloca esta línea aquí

# CONFIGURA AQUÍ
BOT_TOKEN = os.getenv("BOT_TOKEN")
ID_USUARIOS_AUTORIZADOS = [7175478712, 7909467383, 5809993174]
ID_GRUPO_ASESORES = -1002875911448
NOMBRE_CARPETA_DRIVE = "REPORTE_ETIQUETADO_V2"
DRIVE_ID = "1PuCOsjdZZuV0xzAy9ljv5cglORWP1M2n"  # Coloca aquí tu ID de unidad compartida
ALLOWED_CHATS = [-1002640857147, -1002452068425, -4718591093, -4831456255, -1002814603547, -1002838776671, -4951443286, -4870196969, -4824829490, -4979512409, -4903731585, -4910534813, -4845865029, -4643755320, -4860386920, -4945504804, -4854616787, -4979142096, -4653414566]  # Reemplaza con los IDs de tus grupos

def chat_permitido(chat_id: int) -> bool:
    """Verifica si el chat está permitido"""
    return chat_id in ALLOWED_CHATS

# ---- MENSAJES PARA BOT ----
def es_comando_para_bot(update: Update, bot_username: str, comando: str) -> bool:
    """
    Verifica que el comando esté dirigido explícitamente a este bot.
    Ejemplo válido: /ayuda@TuBot
    """
    if not update.message or not update.message.text:
        return False

    texto = update.message.text.strip().lower()
    return texto == f"/{comando} @{bot_username.lower()}"

# VARIABLES
registro_datos = {}
nest_asyncio.apply()
logging.basicConfig(level=logging.INFO)

# Obtener el contenido del JSON desde la variable de entorno
cred_json = os.environ.get("GOOGLE_CREDENTIALS_JSON")
if not cred_json:
    raise ValueError("La variable de entorno GOOGLE_CREDENTIALS_JSON no está definida.")

with open("credentials.json", "w") as f:
    f.write(cred_json)

creds = service_account.Credentials.from_service_account_file("credentials.json")
drive_service = build('drive', 'v3', credentials=creds)

# ---- FUNCIONES DE GOOGLE DRIVE ----
def get_or_create_folder(service, folder_name, parent_id=None):
    # 1. Buscar si la carpeta ya existe
    # Nota: Quitamos 'corpora', 'driveId' e 'includeItemsFromAllDrives' que daban problemas
    query = f"name = '{folder_name}' and mimeType = 'application/vnd.google-apps.folder' and trashed = false"
    
    if parent_id:
        query += f" and '{parent_id}' in parents"

    results = service.files().list(
        q=query,
        fields="files(id, name, webViewLink)",
        # supportsAllDrives=True es seguro dejarlo, por si acaso
        supportsAllDrives=True 
    ).execute()
    
    folders = results.get('files', [])
    
    if folders:
        print(f"✅ Carpeta existente encontrada: {folder_name}")
        return folders[0]['id']
    
    # 2. Si no existe, crearla
    print(f"🔨 Creando carpeta nueva: {folder_name}...")
    metadata = {
        'name': folder_name, 
        'mimeType': 'application/vnd.google-apps.folder'
    }
    
    if parent_id:
        metadata['parents'] = [parent_id]
        
    folder = service.files().create(
        body=metadata, 
        fields='id, webViewLink',
        supportsAllDrives=True
    ).execute()
    
    print(f"✨ Carpeta creada: {folder.get('webViewLink')}")
    return folder['id']

def subir_archivo_excel_grupo(nombre_grupo, archivo_local):
    
    # Crear la carpeta principal (REPORTE_ETIQUETADO)
    carpeta_principal_id = get_or_create_folder(drive_service, NOMBRE_CARPETA_DRIVE, parent_id=DRIVE_ID)
    print(f"DEBUG: Carpeta principal es {carpeta_principal_id}") # Ver en consola

    # Crear la carpeta con el nombre del grupo de Telegram
    carpeta_grupo_id = get_or_create_folder(drive_service, nombre_grupo, parent_id=carpeta_principal_id)
    print(f"DEBUG: Carpeta del grupo '{nombre_grupo}' es {carpeta_grupo_id}")

    media = MediaFileUpload(
        archivo_local,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        resumable=True
    )
    metadata_archivo = {'name': os.path.basename(archivo_local), 'parents': [carpeta_grupo_id]}
    archivo = drive_service.files().create(
        body=metadata_archivo,
        media_body=media,
        fields='id, webViewLink',
        supportsAllDrives=True
    ).execute()
    logging.info(f"✅ Subido {archivo_local} al grupo {nombre_grupo} en Drive.")
    return archivo.get('webViewLink')

def subir_archivos_drive_secuencial():
    if not os.path.exists("reportes"):
        return
    for archivo in os.listdir("reportes"):
        if archivo.endswith(".xlsx"):
            grupo = archivo.split("_")[0]
            ruta_archivo = os.path.join("reportes", archivo)
            try:
                subir_archivo_excel_grupo(grupo, ruta_archivo)
                logging.info(f"✅ Subido: {archivo}")

                # 🚀 Eliminar archivo local tras subirlo
                os.remove(ruta_archivo)
                logging.info(f"🧹 Archivo eliminado: {archivo}")

            except Exception as e:
                logging.error(f"❌ Error al subir {archivo}: {e}")

# ---- FUNCIONES DE EXCEL ----
def crear_directorio_excel():
    if not os.path.exists("reportes"):
        os.makedirs("reportes")

def obtener_nombre_archivo_excel(nombre_grupo):
    fecha_actual = datetime.now(LIMA_TZ).strftime("%Y-%m-%d")
    return f"reportes/{nombre_grupo}_{fecha_actual}.xlsx"

def guardar_en_excel(update, context, datos):
    from io import BytesIO

    nombre_grupo = update.effective_chat.title or f"GRUPO_{update.effective_chat.id}"
    nombre_limpio = re.sub(r'[\\/*?:"<>|]', '_', nombre_grupo.upper().strip())
    fecha_actual = datetime.now(LIMA_TZ).strftime("%Y-%m-%d %H:%M:%S")
    archivo_excel = obtener_nombre_archivo_excel(nombre_limpio)

    if not os.path.exists(archivo_excel):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "FECHA", "CALLE Y CUADRA", "FOTO ANTES", "FOTO DESPUÉS", "FOTO ETIQUETA",
            "LATITUD DEL PUNTO FOTOGRAFIADO", "LONGITUD DEL PUNTO FOTOGRAFIADO"
        ])
        for col in ['F', 'G']:
            ws[f"{col}1"].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        wb.save(archivo_excel)
        print(f"📝 Nuevo archivo Excel creado: {archivo_excel}")

    wb = load_workbook(archivo_excel)
    ws = wb.active
    fila = ws.max_row + 1
    ws.cell(row=fila, column=1, value=fecha_actual)
    ws.cell(row=fila, column=2, value=datos.get("calle_y_cuadra", ""))

    fotos = [datos.get("foto_antes"), datos.get("foto_despues"), datos.get("foto_etiqueta")]
    for idx, ruta in enumerate(fotos, start=3):
        if ruta:
            with open(ruta, 'rb') as f:
                img = PILImage.open(f)
                output = BytesIO()
                img.save(output, format='PNG')
                output.seek(0)
                imagen_excel = ExcelImage(output)
                imagen_excel.width = 150
                imagen_excel.height = 120
                cell_coord = f"{chr(64 + idx)}{fila}"
                ws.add_image(imagen_excel, cell_coord)
                ws.column_dimensions[chr(64 + idx)].width = 25
            ws.row_dimensions[fila].height = 110

    ws.cell(row=fila, column=6, value=datos.get("latitud", ""))
    ws.cell(row=fila, column=7, value=datos.get("longitud", ""))
    wb.save(archivo_excel)
    print(f"✅ Registro agregado al Excel: {archivo_excel}")


# COMUNICACION DRIVE - BOT TELEGRAM

async def test_drive(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id not in ID_USUARIOS_AUTORIZADOS:
        return

    try:
        # 1. Intentar listar archivos (Prueba de comunicación)
        results = drive_service.files().list(
            pageSize=1, 
            fields="files(id, name)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            driveId=DRIVE_ID,
            corpora='drive'
        ).execute()
        
        # 2. Intentar obtener o crear la carpeta principal
        folder_id = get_or_create_folder(drive_service, NOMBRE_CARPETA_DRIVE, parent_id=DRIVE_ID)
        
        await update.message.reply_text(
            f"✅ **Conexión Exitosa**\n"
            f"📂 Carpeta Principal ID: `{folder_id}`\n"
            f"📡 Comunicación con Drive OK.", 
            parse_mode="Markdown"
        )
    except Exception as e:
        await update.message.reply_text(f"❌ **Error de conexión:**\n`{str(e)}`", parse_mode="Markdown")




# ---- NUEVO COMANDO: SUBIR BACKUP ----
async def upload(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not chat_permitido(chat_id):
        return  # Bloquea si el chat no está en ALLOWED_CHATS

    # 👇 Aquí validamos que el mensaje sea para tu bot
    if update.message.chat.type in ['group', 'supergroup']:
        if not es_comando_para_bot(update, context.bot.username, "upload"):
            return

    user_id = update.effective_user.id
    if user_id not in ID_USUARIOS_AUTORIZADOS:
        return await update.message.reply_text("⛔ No tienes permiso para usar este comando.")

    nombre_grupo = update.effective_chat.title or f"GRUPO_{update.effective_chat.id}"
    nombre_limpio = re.sub(r'[\\/*?:"<>|]', '_', nombre_grupo.upper().strip())
    archivo_excel = obtener_nombre_archivo_excel(nombre_limpio)

    if os.path.exists(archivo_excel):
        try:
            enlace = subir_archivo_excel_grupo(nombre_limpio, archivo_excel)
            await update.message.reply_text("☁️ *Carga Exitosa. Hasta mañana*", parse_mode="Markdown")
        except Exception as e:
            await update.message.reply_text(f"❌ Error al subir el archivo")
    else:
        await update.message.reply_text("❌ No hay archivo el día de hoy.")


# COMANDOS
# Funciones principales
async def manejar_no_permitido(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not chat_permitido(chat_id):
        return
        
    await update.message.reply_text("⚠️ Este tipo de mensaje no está permitido en este paso.")

# ✅ Aquí colocas la función de manejo de errores
async def manejar_errores(update: object, context: ContextTypes.DEFAULT_TYPE):
    logging.error(f"❌ Error inesperado: {context.error}")
    
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not chat_permitido(chat_id):
        return  # Bloquea el comando en otros grupos

    # 👇 Aquí validamos que el mensaje sea para tu bot
    if update.message.chat.type in ['group', 'supergroup']:
        if not es_comando_para_bot(update, context.bot.username, "start"):
            return

    user_id = update.effective_user.id
    chat_id = update.effective_chat.id
    registro_datos[(chat_id, user_id)] = {"paso": 0}
    await update.message.reply_text(
        "Hola 👷‍♀️👷‍♂️, comencemos💪💪:\n\n"
        "🔔 Recuerda usar tus EPPs 👷‍♀️👷‍♂️, señalizar 🚧 y evaluar correctamente 🧐 tu zona de trabajo 🧰\n\n"
        "✍️ Ahora: Escribe el nombre de la calle y cuadra a intervenir\n\n"
        "📌Ejm: Av. Bolívar - Cdra 5\n"
        "📌Ejm: Ca. Leoncio Prado - Mz. B1\n"
        "📌Ejm: Psje. SN - S/N\n\n"
        "¡Vamos, tú puedes!💪"
    )

async def ayuda(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not chat_permitido(chat_id):
        return
        
    # 👇 Aquí validamos que el mensaje sea para tu bot
    if update.message.chat.type in ['group', 'supergroup']:
        if not es_comando_para_bot(update, context.bot.username, "ayuda"):
            return

    botones = [
        [InlineKeyboardButton("🔄 Reiniciar registro", callback_data="reiniciar")],
        [InlineKeyboardButton("🔎 ¿En qué paso estoy?", callback_data="ver_paso")],
        [InlineKeyboardButton("🆘 Solicitar ayuda de un asesor", callback_data="asesor")],
    ]

    texto = (
        "🧰 *Centro de ayuda Equipo Ordenamiento* 👷‍♂️👷‍♀️\n\n"
        "Selecciona una de las siguientes opciones para continuar tu trabajo de forma asistida:\n\n"
        "🔄 *Reiniciar registro*: Si deseas comenzar de nuevo.\n\n"
        "🆘 *Solicitar ayuda de un asesor*: Si necesitas asistencia inmediata en este momento.\n\n"
        "🔎 *¿En qué paso estoy?*: Te recordaremos en qué parte del registro te quedaste.\n\n"
        "💡 *Consejo del día:* _Una buena señalización 🚧 y una foto clara valen más que mil palabras.💪_\n\n"
    )

    await update.message.reply_text(texto, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones))

async def callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not chat_permitido(chat_id):
        return
        
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    chat_id = query.message.chat.id

    if query.data == "reiniciar":
        registro_datos[(chat_id, user_id)] = {"paso": 0}
        await query.edit_message_text(
        "🔄 *Registro reiniciado correctamente*.\n\n"
        "✍️ Por favor, envía el nombre de la calle y cuadra a intervenir para comenzar de nuevo.\n"
        "📌 Ejm: Av. Grau - Cdra 12\n\n"
        "¡Tranquilo, Vamos de nuevo! 💪",
        parse_mode="Markdown"
    )
    elif query.data == "asesor":
        nombre = query.from_user.full_name
        username = query.from_user.username
        fecha_hora = datetime.now(LIMA_TZ).strftime("%d/%m/%Y %H:%M")
        mensaje = f"📢 Se ha solicitado ayuda de un asesor:\n👤 Usuario: {nombre} (@{username})\n🆔 Grupo: {chat_id} ({update.effective_chat.title})\n⏰ Hora: {fecha_hora}"
        await context.bot.send_message(chat_id=ID_GRUPO_ASESORES, text=mensaje)
        await query.edit_message_text("✅ Se ha notificado a un asesor 👨‍💻👨‍💻. Te contactarán en un momento.")
    elif query.data == "ver_paso":
        datos = registro_datos.setdefault((chat_id, user_id), {"paso": 0})
        paso = datos.get("paso", 0)
        
        mensajes = {
        0: "✍️ *Te encuentras en el Inicio del registro.*\n\n👉 Escribe el *nombre de la calle y cuadra* a intervenir.\n\n📌 Ejm: Av. Grau - Cdra 12\n",
        1: "📸 *Estás en el Paso de la foto del ANTES.*\n\n👉 Envía una *foto clara* del punto antes de intervenir.\n\n🔔 Recuerda que las fotos se toman de manera vertical🧐\n",
        2: "📸 *Estás en el Paso de la foto del DESPUÉS.*\n\n👉 Envía una *foto clara* del punto tras la intervención.\n\n🔔 Recuerda que las fotos se toman de manera vertical🧐\n",
        3: "🏷️ *Estás en el Paso de la foto de la ETIQUETA.*\n\n👉 Toma una foto que muestre claramente la etiqueta de instalación.\n\n🔔 Recuerda que las fotos se toman de manera vertical🧐\n",
        4: "📍 *Estás en el último paso: ubicación GPS.*\n\n👉 Comparte tu ubicación GPS actual desde donde tomaste las fotos.",
        }

    texto = (
        f"🔎 *Estado actual del registro* 👷‍♂️👷‍♀️\n\n"
        f"{mensajes.get(paso, '⚠️ Estado desconocido.')}\n\n"
        f"🔔 *Recuerda:* responde este mensaje con el contenido solicitado para continuar correctamente. 💪"
    )

    await query.edit_message_text(texto, parse_mode="Markdown")

# FLUJO REGISTRO
async def manejar_texto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not chat_permitido(chat_id):
        return
        
    if update.message.chat.type in ['group', 'supergroup']:
        if not (update.message.reply_to_message and update.message.reply_to_message.from_user.id == context.bot.id):
            return

    user_id = update.effective_user.id
    chat_id = update.effective_chat.id
    datos = registro_datos.setdefault((chat_id, user_id), {})

    if datos.get("paso") != 0:
        await update.message.reply_text("⚠️ Este paso no requiere texto. Usa el botón adecuado o responde con el tipo correcto.")
        return

    datos['calle_y_cuadra'] = update.message.text
    datos["paso"] = 1
    botones = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔁 Repetir este paso 📝", callback_data="repetir_paso_0")],
        [InlineKeyboardButton("➡️ Continuar con fotos del ANTES", callback_data="continuar_paso_1")]
    ])

    await update.message.reply_text(
            "✅ ¡Excelente! Ya tengo el nombre de la calle y cuadra 🛣️\n\n"
            "📸 Ahora toca tomar foto del antes 💪.\n\n"
            "📲 Recuerda usar el *AppNoteCam* para la toma de fotos📸\n\n"
            "🔔 Recuerda que las fotos se toman de manera vertical🧐\n\n"
            "Presiona *Continuar* para seguir con la foto del ANTES 💪 o *Repetir* si deseas hacer alguna correción. 🧐\n\n"
            "👉 Cuando estés listo, selecciona una opción:",
            parse_mode="Markdown",
            reply_markup=botones
    )

async def manejar_foto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not chat_permitido(chat_id):
        return
        
    if update.message.chat.type in ['group', 'supergroup']:
        if not (update.message.reply_to_message and update.message.reply_to_message.from_user.id == context.bot.id):
            return

    user_id = update.effective_user.id
    chat_id = update.effective_chat.id
    datos = registro_datos.setdefault((chat_id, user_id), {})
    paso = datos.get("paso", 0)

    if paso not in [1, 2, 3]:
        await update.message.reply_text("⚠️ Este paso no requiere fotos. Usa el botón adecuado o responde con lo solicitado.")
        return

    archivo = await update.message.photo[-1].get_file()

    if paso == 1:
        ruta = f"reportes/{chat_id}_{user_id}_antes.jpg"
        await archivo.download_to_drive(ruta)
        datos['foto_antes'] = ruta
        datos["paso"] = 2
        botones = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔁 Repetir esta foto 📸", callback_data="repetir_paso_1")],
            [InlineKeyboardButton("➡️ Continuar con foto del DESPUÉS", callback_data="continuar_paso_2")]
        ])
        await update.message.reply_text(
            "📸 ¡Genial! Ya tengo la foto del ANTES 👀\n\n"
            "🎯 Ahora necesito que tomes la foto del DESPUÉS 📸\n\n"
            "📲 Recuerda usar el *AppNoteCam* para la toma de fotos📸\n\n"
            "🔔 Recuerda que las fotos se toman de manera vertical🧐\n\n"
            "Presiona *Continuar* para seguir con la foto del DESPUÉS 💪 o *Repetir* si deseas hacer alguna correción. 🧐",
            parse_mode="Markdown",
            reply_markup=botones
        )
    elif paso == 2:
        ruta = f"reportes/{chat_id}_{user_id}_despues.jpg"
        await archivo.download_to_drive(ruta)
        datos['foto_despues'] = ruta
        datos["paso"] = 3
        botones = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔁 Repetir esta foto 📸", callback_data="repetir_paso_2")],
            [InlineKeyboardButton("➡️ Continuar con foto de ETIQUETA", callback_data="continuar_paso_3")]
        ])
        await update.message.reply_text(
            "📷 ¡Excelente trabajo! Ya tengo la foto del DESPUÉS ✅\n\n"
            "🔖 Ahora toca foto de la ETIQUETA que dejaste en la fibra ADSS.\n\n"
            "📲 Recuerda usar el *AppNoteCam* para la toma de fotos📸\n\n"
            "📷 La foto de la etiqueta debe ser de cerca y con el rotulo que escribiste legible ✍️ 🧐\n\n"
            "🔔 Recuerda que las fotos se toman de manera vertical🧐\n\n"
            "Presiona *Continuar* para seguir con la foto del la ETIQUETA 💪 o *Repetir* si deseas hacer alguna correción. 🧐",
            parse_mode="Markdown",
            reply_markup=botones
        )

    elif paso == 3:
        ruta = f"reportes/{chat_id}_{user_id}_etiqueta.jpg"
        await archivo.download_to_drive(ruta)
        datos['foto_etiqueta'] = ruta
        datos["paso"] = 4
        botones = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔁 Repetir esta foto 🏷️", callback_data="repetir_paso_3")],
            [InlineKeyboardButton("➡️ Continuar con la ubicación GPS 📍", callback_data="continuar_paso_4")]
        ])
        await update.message.reply_text(
            "🏷️ ¡Foto de la ETIQUETA recibida, excelente trabajo! 📌\n\n"
            "🧭 Ahora necesitamos tu *Ubicación GPS* exacta del punto intervenido.\n"
            "Presiona *Continuar* para compartir tu Ubicación GPS 💪 o *Repetir* si deseas hacer alguna correción. 🧐",
            parse_mode="Markdown",
            reply_markup=botones
        )

async def manejar_ubicacion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not chat_permitido(chat_id):
        return
        
    if update.message.chat.type in ['group', 'supergroup']:
        if not (update.message.reply_to_message and update.message.reply_to_message.from_user.id == context.bot.id):
            return

    user_id = update.effective_user.id
    chat_id = update.effective_chat.id
    datos = registro_datos.get((chat_id, user_id))

    if not datos or datos.get("paso") != 4:
        await update.message.reply_text("⚠️ Este paso requiere tu ubicación GPS. Usa el botón adecuado o responde con tu ubicación GPS actual.")
        return

    if all(k in datos for k in ['foto_antes', 'foto_despues', 'foto_etiqueta']):
        datos['latitud'] = update.message.location.latitude
        datos['longitud'] = update.message.location.longitude
        loop = asyncio.get_running_loop()
        await loop.run_in_executor(None, guardar_en_excel, update, context,datos)
        await update.message.reply_text(
            "✅ ¡Registro completado con éxito! 🎯\n\n"
            "📊 Se ha guardado la informacion y fotos compartidas correctamente.\n\n"
            "🔁 Si deseas iniciar otro registro, simplemente escribe /start y etiquetame, para poder ayudarte 💪💪\n\n"
            "¡Sigue así! 🦾"
        )
        del registro_datos[(chat_id, user_id)]
    else:
        await update.message.reply_text("📷 Por favor, completa primero el envío de todas las fotos requeridas.")


async def exportar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not chat_permitido(chat_id):
        return

    # 👇 Aquí validamos que el mensaje sea para tu bot
    if update.message.chat.type in ['group', 'supergroup']:
        if not es_comando_para_bot(update, context.bot.username, "exportar"):
            return
    
    user_id = update.effective_user.id
    chat = update.effective_chat
    if user_id not in ID_USUARIOS_AUTORIZADOS:
        return await update.message.reply_text("⛔ No tienes permiso para usar este comando.")
    
    nombre_grupo = update.effective_chat.title or f"GRUPO_{chat_id}"
    nombre_limpio = re.sub(r'[\\/*?:"<>|]', '_', nombre_grupo.upper().strip())
    nombre_archivo = obtener_nombre_archivo_excel(nombre_limpio)

    if os.path.exists(nombre_archivo):
        await context.bot.send_document(chat_id=chat.id, document=open(nombre_archivo, "rb"))
    else:
        await update.message.reply_text("❌ No hay registros para exportar hoy.")

async def manejo_navegacion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not chat_permitido(chat_id):
        return
        
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    chat_id = query.message.chat.id
    datos = registro_datos.setdefault((chat_id, user_id), {})

    paso = datos.get("paso", 0)
    data = query.data

    if data == "repetir_paso_0":
        datos.pop("calle_y_cuadra", None)
        datos["paso"] = 0
        await query.edit_message_text("🔁 Reiniciado. Por favor, digita ✍️ el nombre de la calle y cuadra.\n\n📌Ejm: Av. Bolívar - Cdra 5\n📌Ejm: Ca. Leoncio Prado - Mz. B1\n📌Ejm: Psje. SN - S/N")

    elif data == "continuar_paso_1":
        datos["paso"] = 1
        await query.edit_message_text("🖼️ Envía la foto del ANTES.\n\n🔔 Recuerda que las fotos se toman de manera vertical🧐")

    elif data == "repetir_paso_1":
        datos.pop("foto_antes", None)
        datos["paso"] = 1
        await query.edit_message_text("🔁 Reiniciado. Por favor, envía la foto del ANTES.\n\n🔔 Recuerda que las fotos se toman de manera vertical🧐")

    elif data == "continuar_paso_2":
        datos["paso"] = 2
        await query.edit_message_text("🖼️ Envía la foto del DESPUÉS.\n\n🔔 Recuerda que las fotos se toman de manera vertical🧐")

    elif data == "repetir_paso_2":
        datos.pop("foto_despues", None)
        datos["paso"] = 2
        await query.edit_message_text("🔁 Reiniciado. Por favor, envía la foto del DESPUÉS.\n\n🔔 Recuerda que las fotos se toman de manera vertical🧐")

    elif data == "continuar_paso_3":
        datos["paso"] = 3
        await query.edit_message_text("🏷️ Envía la foto de la ETIQUETA.\n\n🔔 Recuerda que las fotos se toman de manera vertical🧐")

    elif data == "repetir_paso_3":
        datos.pop("foto_etiqueta", None)
        datos["paso"] = 3
        await query.edit_message_text("🔁 Reiniciado. Por favor, envía la foto de la ETIQUETA.\n\n🔔 Recuerda que las fotos se toman de manera vertical🧐")

    elif data == "continuar_paso_4":
        datos["paso"] = 4
        await query.edit_message_text("📍 Comparte tu ubicación GPS actual.\n\n🔔 Recuerda que debe de ser del punto intervenido 🧐")

    elif data == "repetir_paso_4":
        datos.pop("latitud", None)
        datos.pop("longitud", None)
        datos["paso"] = 4
        await query.edit_message_text("🔁 Reiniciado. Por favor, vuelve a enviar tu ubicación GPS actual.\n\n🔔 Recuerda que debe de ser del punto intervenido 🧐")

async def manejar_no_permitido(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.chat.type in ['group', 'supergroup']:
        if not (update.message.reply_to_message and update.message.reply_to_message.from_user.id == context.bot.id):
            return

    user_id = update.effective_user.id
    chat_id = update.effective_chat.id
    datos = registro_datos.get((chat_id, user_id), {})
    paso = datos.get("paso", None)

    mensajes = {
        0: "❌ Este paso solo requiere texto (nombre de calle y cuadra).❌",
        1: "❌ Este paso solo requiere una foto del ANTES.❌",
        2: "❌ Este paso solo requiere una foto del DESPUÉS.❌",
        3: "❌ Este paso solo requiere una foto de la ETIQUETA.❌",
        4: "❌ Este paso solo requiere tu ubicación GPS.❌",
        None: "❌ Aún no has iniciado el registro. Usa /start para comenzar."
    }

    await update.message.reply_text(mensajes.get(paso, "❌ Este contenido no es válido para este paso.❌"))

async def get_chat_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not chat_permitido(chat_id):
        return

    # 👇 Aquí validamos que el mensaje sea para tu bot
    if update.message.chat.type in ['group', 'supergroup']:
        if not es_comando_para_bot(update, context.bot.username, "id"):
            return
    
    await update.message.reply_text(f"Chat ID: {update.effective_chat.id}")

# ---- MAIN ----
async def main():
    crear_directorio_excel()
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("testdrive", test_drive))
    app.add_handler(CommandHandler("ayuda", ayuda))
    app.add_handler(CommandHandler("exportar", exportar))
    app.add_handler(CommandHandler("upload", upload))
    app.add_handler(CommandHandler("id", get_chat_id))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, manejar_texto))
    app.add_handler(MessageHandler(filters.PHOTO, manejar_foto))
    app.add_handler(MessageHandler(filters.LOCATION, manejar_ubicacion))
    app.add_handler(CallbackQueryHandler(manejo_navegacion, pattern=r"^(repetir_paso_|continuar_paso_)"))
    app.add_handler(MessageHandler(~filters.TEXT & ~filters.PHOTO & ~filters.LOCATION, manejar_no_permitido))
    app.add_handler(CallbackQueryHandler(callback_handler))
    app.add_error_handler(manejar_errores)

    # Programar subida diaria a las 18:30 hora Perú
    scheduler = AsyncIOScheduler()
    scheduler.add_job(
        subir_archivos_drive_secuencial,
        'cron',
        hour=12,
        minute=12,
        timezone=timezone('America/Lima')
    )
    scheduler.start()

    await app.bot.delete_webhook(drop_pending_updates=True)
    await app.run_polling()

if __name__ == "__main__":
    import asyncio
    nest_asyncio.apply()
    asyncio.get_event_loop().run_until_complete(main())
